"""
Python port of MVsourceFunctions.R

Ported as faithfully as possible to R semantics (including R's specific
strsplit() behaviour, 1-indexing translated to 0-indexing, and paste()'s
NA -> "NA" string coercion). See PORTING_NOTES.md for caveats.
"""
import math
import string
from collections import OrderedDict
import pandas as pd
import numpy as np
import openpyxl

LETTERS = string.ascii_uppercase


class MVError(Exception):
    """Raised for any input/processing error, mirrors R's errMessage flow."""
    pass


# ---------------------------------------------------------------------------
# small helpers that replicate R-specific behaviour
# ---------------------------------------------------------------------------

def r_strsplit(s, sep=' '):
    """Mimic R's strsplit(s, sep, fixed=TRUE) for a single string:
    - leading delimiter -> leading "" element (kept)
    - trailing delimiter -> trailing "" element (DROPPED)
    - empty string input -> [] (R: character(0))
    """
    if s is None:
        return []
    s = str(s)
    if s == '':
        return []
    parts = s.split(sep)
    if s.endswith(sep) and len(parts) > 0:
        parts = parts[:-1]
    return parts


def r_get(parts, idx):
    """1-indexed-style safe get -> None if out of range (R: NA)."""
    return parts[idx] if 0 <= idx < len(parts) else None


def rpaste(*args, sep=' '):
    """R's paste(): NA (None) arguments are coerced to the string 'NA'."""
    return sep.join('NA' if a is None else str(a) for a in args)


def to_num(x):
    """gsub(',', '.', x) %>% as.numeric() ; None -> NaN"""
    if x is None:
        return float('nan')
    if isinstance(x, (int, float)):
        return float(x)
    x = str(x).replace(',', '.').strip()
    if x == '' or x.upper() == 'NA':
        return float('nan')
    return float(x)


# ---------------------------------------------------------------------------
# READ INPUT
# ---------------------------------------------------------------------------

def get_stock_list(file_name):
    """C1:M2 -> {drug_name: conc(float)}, dropping all-NA columns."""
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb[wb.sheetnames[0]]
    names, values = [], []
    for col in range(3, 14):  # C=3 .. M=13
        name = ws.cell(row=1, column=col).value
        val = ws.cell(row=2, column=col).value
        if name is None and val is None:
            continue
        names.append(name)
        values.append(val)
    stock_list = OrderedDict()
    for n, v in zip(names, values):
        stock_list[n] = to_num(v)
    return stock_list


def get_well_vols(file_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb[wb.sheetnames[0]]
    total_vol = ws['C5'].value
    fill_vol = ws['C6'].value
    return {'TotalVol': float(total_vol), 'FillVol': float(fill_vol)}


def get_n_plate(file_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return int(ws['F6'].value)


def get_plate_map(file_name):
    """B57:M64 (8 rows x 12 cols) -> DataFrame[Well, fillID, solID,
    DrugType, DrugConc, Solvent, Inoc]"""
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for r in range(8):  # rows A..H -> excel rows 57..64
        excel_row = 57 + r
        for c in range(12):  # cols 1..12 -> excel cols B..M (2..13)
            excel_col = 2 + c
            raw = ws.cell(row=excel_row, column=excel_col).value
            well_id = f"{LETTERS[r]}{c + 1}"
            if raw is None:
                raw = ''
            # readxl's read_xlsx() trims leading/trailing whitespace by
            # default (trim_ws=TRUE) -- this matters a lot here, since the
            # border/no-drug well formula produces e.g. " 0 MED_A " with a
            # leading space that must be gone before the "is this a FILL
            # well" check below works.
            raw = str(raw).strip()
            if raw == '0' or raw == '':
                continue

            parsed = r_strsplit(raw, ' ')
            p0 = r_get(parsed, 0)

            if p0 == '0':
                # "blank fill well" branch -- kept for fidelity, but in
                # practice this template's blank-well strings never hit it
                # (see PORTING_NOTES.md)
                solid = 'FILL'
                drugtype = 'NA'
                conc = r_get(parsed, 0)
                solvent = r_get(parsed, 1)
                inoc = 'NA'
            else:
                p1 = r_get(parsed, 1)
                p2 = r_get(parsed, 2)
                p3 = r_get(parsed, 3)
                solid = rpaste(p0, p1, p2)
                drugtype = p0
                conc = p1
                solvent = p2
                inoc = p3 if p3 is not None else 'NA'

            rows.append({
                'Well': well_id, 'fillID': raw, 'solID': solid,
                'DrugType': drugtype, 'DrugConc': conc,
                'Solvent': solvent, 'Inoc': inoc,
            })

    fin_map = pd.DataFrame(rows)
    fin_map['DrugConc'] = fin_map['DrugConc'].apply(to_num)
    return fin_map


# ---------------------------------------------------------------------------
# PREPARATION
# ---------------------------------------------------------------------------

def create_sol_list(plate_map, total_vol_well, inoc_vol, stock_list, n_plate):
    plate_map = plate_map.copy()
    plate_map['solID'] = plate_map['solID'].astype(str).str.replace(',', '.', regex=False)

    occ = plate_map['solID'].value_counts()
    occurences = pd.DataFrame({'solID': occ.index, 'Occ': occ.values * float(n_plate)})

    fin_list = plate_map[['solID', 'DrugType', 'DrugConc', 'Solvent']].drop_duplicates()
    fin_list = fin_list.merge(occurences, on='solID', how='left')
    fin_list = fin_list[fin_list['solID'] != 'FILL'].reset_index(drop=True)
    fin_list.columns = ['SolID', 'DrugType', 'DrugConc', 'Solvent', 'Occurence']

    fin_list['Occurence'] = fin_list['Occurence'].astype(float)
    fin_list['DrugConc'] = fin_list['DrugConc'].astype(float)

    fin_list = calculate_dil_volume(fin_list, total_vol_well, inoc_vol, stock_list)
    return fin_list


def calculate_dil_volume(sol_list, total_vol_well, inoc_vol, stock_list):
    sol_list = sol_list.copy()
    drug_sol_well = total_vol_well - inoc_vol
    sol_list['solAmt'] = sol_list['Occurence'] * drug_sol_well + 150  # 150 uL excess (per R comment "10uL" is stale)

    sol_list['DrugConc'] = sol_list['DrugConc'].astype(float) * total_vol_well / (total_vol_well - inoc_vol)

    # remove no-drug (border/blank) solutions
    sol_list = sol_list[sol_list['DrugType'] != ''].reset_index(drop=True)
    sol_list = sol_list[sol_list['DrugType'].notna()].reset_index(drop=True)

    new_rows = []
    solvents = sol_list['Solvent'].unique().tolist()
    drugs = sol_list['DrugType'].unique().tolist()

    for solv in solvents:
        for drug in drugs:
            cur_list = sol_list[(sol_list['DrugType'] == drug) & (sol_list['Solvent'] == solv)].copy()
            if len(cur_list) == 0:
                continue
            cur_list['DrugConc'] = cur_list['DrugConc'].astype(float)
            cur_list = cur_list.sort_values('DrugConc').reset_index(drop=True)

            # ---- stage 1: insert pre-dilutions where dilution factor > 10 ----
            new_cur_rows = []
            n = len(cur_list)
            for q in range(n):
                row_q = cur_list.iloc[q].to_dict()
                new_cur_rows.append(dict(row_q))

                if q == n - 1:
                    conc_hi = stock_list[cur_list['DrugType'].iloc[q]]
                else:
                    conc_hi = cur_list['DrugConc'].iloc[q + 1]
                # guard mirrors R: DrugConc==0 (no-drug control) is
                # excluded before the ratio is ever used, but compute it
                # safely either way to avoid a spurious /0 warning
                q_conc = cur_list['DrugConc'].iloc[q]
                cur_dil_fac = (conc_hi / q_conc) if q_conc != 0 else float('inf')

                if q_conc > 0 and cur_dil_fac > 10:
                    nex = dict(row_q)
                    while cur_dil_fac > 10:
                        nex = dict(nex)
                        nex['DrugConc'] = conc_hi / 10
                        nex['Occurence'] = 0
                        nex['SolID'] = f"{nex['DrugType']} {nex['DrugConc']} {nex['Solvent']}"
                        new_cur_rows.append(dict(nex))
                        conc_hi = nex['DrugConc']
                        cur_dil_fac = conc_hi / cur_list['DrugConc'].iloc[q]

            cur_list2 = pd.DataFrame(new_cur_rows)
            cur_list2['solAmt'] = cur_list2['solAmt'].astype(float)
            cur_list2['DrugConc'] = cur_list2['DrugConc'].astype(float)
            cur_list2 = cur_list2.sort_values('DrugConc').reset_index(drop=True)

            # ---- stage 2: amounts needed from the tube above ----
            needed_from_above = []
            m_n = len(cur_list2)
            for m in range(m_n):
                if m < m_n - 1:
                    amt_needed = cur_list2['solAmt'].iloc[m] * cur_list2['DrugConc'].iloc[m] / cur_list2['DrugConc'].iloc[m + 1]
                    if 0 < amt_needed < 30:
                        cur_list2.loc[m, 'solAmt'] = cur_list2['solAmt'].iloc[m] * 30 / amt_needed
                        amt_needed = 30
                    needed_from_above.append(amt_needed)
                    cur_list2.loc[m + 1, 'solAmt'] = cur_list2['solAmt'].iloc[m + 1] + amt_needed
                else:
                    stock_conc = stock_list[cur_list2['DrugType'].iloc[m]]
                    amt_needed = cur_list2['solAmt'].iloc[m] * cur_list2['DrugConc'].iloc[m] / stock_conc
                    if 0 < amt_needed < 30:
                        cur_list2.loc[m, 'solAmt'] = cur_list2['solAmt'].iloc[m] * 30 / amt_needed
                        amt_needed = 30
                    needed_from_above.append(amt_needed)

            cur_list2['AmtHi'] = needed_from_above
            new_rows.append(cur_list2)

    if not new_rows:
        raise MVError("No drug solutions found to build a dilution scheme from.")

    new_sol_list = pd.concat(new_rows, ignore_index=True)
    new_sol_list['solventAmt'] = new_sol_list['solAmt'].astype(float) - new_sol_list['AmtHi'].astype(float)
    new_sol_list['reqTube'] = '15_Falcon'

    if new_sol_list['solAmt'].astype(float).max() >= 14 * 1000:
        raise MVError("OVER CAPACITY!")

    return new_sol_list


def create_dil_map(sol_list, deck_map, stock_list):
    """deck_map: OrderedDict labware_N -> description string"""
    rows = []
    for key, desc in deck_map.items():
        if 'Falcon' in desc:
            for i in range(3):
                for j in range(1, 6):
                    rows.append({'Slot': f"{LETTERS[i]}{j}", 'Fill': '', 'Labware': key, 'solutionType': ''})
    solution_map = pd.DataFrame(rows)

    stock_labware = [k for k, v in deck_map.items() if 'stock' in v][0]
    stock_idx = solution_map.index[solution_map['Labware'] == stock_labware].tolist()
    stock_names = list(stock_list.keys())
    for idx, name in zip(stock_idx[:len(stock_names)], stock_names):
        solution_map.loc[idx, 'Fill'] = name
        solution_map.loc[idx, 'solutionType'] = 'Stock'

    solution_map = solution_map[solution_map['solutionType'] != 'Stock'].reset_index(drop=True)

    if len(sol_list) > len(solution_map):
        raise MVError("Too many solution types!")

    sol_ids = sol_list['SolID'].tolist()
    for i, sid in enumerate(sol_ids):
        solution_map.loc[i, 'Fill'] = sid

    solution_map = solution_map[['Slot', 'Fill', 'Labware']]
    solution_map = solution_map[solution_map['Fill'] != ''].reset_index(drop=True)
    return solution_map


# ---------------------------------------------------------------------------
# COMMANDS
# ---------------------------------------------------------------------------

CMD_COLS = ['SourceLabware', 'SourceSlot', 'TargetLabware', 'TargetSlot',
            'TransAmt', 'MixAmt', 'TipID', 'Comment']


def _name_for_value(deck_map, value):
    for k, v in deck_map.items():
        if v == value:
            return k
    return None


def cmd_init_dist(deck_map, sol_list, solvent_map, dil_map):
    rows = []
    tip_id = 1
    solvents = sol_list['Solvent'].unique().tolist()
    for solv in solvents:
        cur = sol_list[sol_list['Solvent'] == str(solv)]
        for _, r in cur.iterrows():
            src_ware = _name_for_value(deck_map, 'Solvent')
            src_slot = solvent_map.loc[solvent_map['Name'] == str(r['Solvent']), 'Slot'].iloc[0]
            tgt_ware = dil_map.loc[dil_map['Fill'] == str(r['SolID']), 'Labware'].iloc[0]
            tgt_slot = dil_map.loc[dil_map['Fill'] == str(r['SolID']), 'Slot'].iloc[0]
            rows.append([src_ware, src_slot, tgt_ware, tgt_slot,
                         r['solventAmt'], '0', tip_id, 'Initial solvent distribution'])
        tip_id += 1
    return pd.DataFrame(rows, columns=CMD_COLS)


def cmd_hi_drug(cmd_list, sol_list, stock_map, deck_map, dil_map):
    tip_id = int(cmd_list['TipID'].iloc[-1]) + 1 if len(cmd_list) else 1
    rows = []
    for _, srow in stock_map.iterrows():
        cr_sol_list = sol_list[sol_list['DrugType'] == srow['Name']]
        solvents = cr_sol_list['Solvent'].unique().tolist()
        for solv in solvents:
            cur = cr_sol_list[cr_sol_list['Solvent'] == solv]
            cur = cur[cur['DrugConc'].astype(float) == cur['DrugConc'].astype(float).max()]
            r = cur.iloc[0]
            src_ware = _name_for_value(deck_map, '15ml_Falcon_stock')
            src_slot = stock_map.loc[stock_map['Name'] == r['DrugType'], 'Slot'].iloc[0]
            tgt_ware = dil_map.loc[dil_map['Fill'] == r['SolID'], 'Labware'].iloc[0]
            tgt_slot = dil_map.loc[dil_map['Fill'] == r['SolID'], 'Slot'].iloc[0]
            rows.append([src_ware, src_slot, tgt_ware, tgt_slot,
                         r['AmtHi'], r['AmtHi'], tip_id, 'Initial stock dilution'])
            tip_id += 1
    new_rows = pd.DataFrame(rows, columns=CMD_COLS)
    return pd.concat([cmd_list, new_rows], ignore_index=True)


def cmd_serial_dil(cmd_list, sol_list, dil_map):
    tip_id = int(cmd_list['TipID'].astype(float).max()) + 1
    rows = []
    drugs = sol_list['DrugType'].unique().tolist()
    solvents = sol_list['Solvent'].unique().tolist()
    for drug in drugs:
        for solv in solvents:
            cur = sol_list[(sol_list['DrugType'] == drug) & (sol_list['Solvent'] == solv)]
            if len(cur) == 0:
                continue
            cur = cur.sort_values('DrugConc', ascending=False).reset_index(drop=True)
            for m in range(len(cur) - 1):
                amt = cur['AmtHi'].iloc[m + 1]
                if amt != 0:
                    src_ware = dil_map.loc[dil_map['Fill'] == cur['SolID'].iloc[m], 'Labware'].iloc[0]
                    src_slot = dil_map.loc[dil_map['Fill'] == cur['SolID'].iloc[m], 'Slot'].iloc[0]
                    tgt_ware = dil_map.loc[dil_map['Fill'] == cur['SolID'].iloc[m + 1], 'Labware'].iloc[0]
                    tgt_slot = dil_map.loc[dil_map['Fill'] == cur['SolID'].iloc[m + 1], 'Slot'].iloc[0]
                    rows.append([src_ware, src_slot, tgt_ware, tgt_slot, amt, amt,
                                 tip_id, f"Serially diluting to  {cur['SolID'].iloc[m + 1]}"])
                    tip_id += 1
    new_rows = pd.DataFrame(rows, columns=CMD_COLS)
    return pd.concat([cmd_list, new_rows], ignore_index=True)


def cmd_drug_sol_dist(cmd_list, dil_map, plate_map, deck_map, well_info, n_plates):
    tip_id = int(cmd_list['TipID'].astype(float).max()) + 1
    trans_v = well_info['TotalVol'] - well_info['FillVol']

    plate_map = plate_map.copy()
    plate_map['solID'] = plate_map['solID'].astype(str).str.replace(',', '.', regex=False)
    dil_map = dil_map.copy()
    dil_map['Fill'] = dil_map['Fill'].astype(str).str.replace(',', '.', regex=False)

    plates = [f"96-well_{LETTERS[i]}" for i in range(n_plates)]

    rows = []
    for _, d in dil_map.iterrows():
        target_wells = plate_map.loc[plate_map['solID'] == d['Fill'], 'Well'].tolist()
        target_wells_str = ', '.join(target_wells)
        for p in plates:
            tgt_ware = _name_for_value(deck_map, p)
            rows.append([d['Labware'], d['Slot'], tgt_ware, target_wells_str,
                         trans_v, 0, tip_id, f"Distributing  {d['Fill']}"])
        tip_id += 1
    new_rows = pd.DataFrame(rows, columns=CMD_COLS)
    return pd.concat([cmd_list, new_rows], ignore_index=True)


def cmd_fill_outer(plate_map, deck_map, solvent_map, well_info, cmd_list, n_plates):
    tip_id = int(cmd_list['TipID'].astype(float).max()) + 1
    solvent_map = solvent_map.copy()
    solvent_map['Name'] = solvent_map['Name'].apply(lambda x: 'WATER' if str(x).lower() == 'water' else x)

    cur_plate_map = plate_map[plate_map['solID'] == 'FILL']
    plates = [f"96-well_{LETTERS[i]}" for i in range(n_plates)]

    rows = []
    solvents = solvent_map['Name'].unique().tolist()
    for solv in solvents:
        target_wells = ', '.join(cur_plate_map.loc[cur_plate_map['Solvent'] == solv, 'Well'].tolist())
        for p in plates:
            src_slot = solvent_map.loc[solvent_map['Name'] == solv, 'Slot'].iloc[0]
            tgt_ware = _name_for_value(deck_map, p)
            rows.append([_name_for_value(deck_map, 'Solvent'), src_slot, tgt_ware, target_wells,
                         well_info['TotalVol'], 0, tip_id, 'Filling outer wells with WATER'])
            tip_id += 1
    if rows:
        new_rows = pd.DataFrame(rows, columns=CMD_COLS)
        cmd_list = pd.concat([cmd_list, new_rows], ignore_index=True)
    return cmd_list


def cmd_separate_long(cmd_list):
    """Splits any command targeting >8 wells into <=8-well chunks, each
    getting its own TipID. Critically, every row *after* the one being
    split must have its TipID shifted by however many extra chunks were
    just inserted -- this cascades down the whole list, exactly mirroring
    the R version's in-place update of cmd_list[i:nrow,7]."""
    cmd_list = cmd_list.copy().reset_index(drop=True)
    cmd_list['TipID'] = cmd_list['TipID'].astype(float)
    new_rows = []
    for i in range(len(cmd_list)):
        row = cmd_list.iloc[i]
        rem_wells = [w for w in row['TargetSlot'].split(', ') if w != ''] if row['TargetSlot'] else []
        d_tip = 0
        while rem_wells:
            n_wells = min(8, len(rem_wells))
            nex_wells = rem_wells[:n_wells]
            rem_wells = rem_wells[n_wells:]
            nex = row.copy()
            nex['TargetSlot'] = ', '.join(nex_wells)
            nex['TipID'] = float(row['TipID']) + d_tip
            d_tip += 1
            new_rows.append(nex)
        if d_tip > 0:
            cmd_list.loc[i:, 'TipID'] = cmd_list.loc[i:, 'TipID'] + (d_tip - 1)
    return pd.DataFrame(new_rows, columns=CMD_COLS).reset_index(drop=True)


# ---------------------------------------------------------------------------
# COUNTERS
# ---------------------------------------------------------------------------

def cal_sol_amt(deck_map, solvent_map, cmd_list):
    rack_position = _name_for_value(deck_map, 'Solvent')
    req_amt = []
    for _, r in solvent_map.iterrows():
        rel = cmd_list[(cmd_list['SourceLabware'] == rack_position) & (cmd_list['SourceSlot'] == r['Slot'])]
        n_well = rel['TargetSlot'].apply(lambda x: len([w for w in x.split(', ') if w != '']))
        solvent_amt = (rel['TransAmt'].astype(float) * n_well).sum()
        req_amt.append(solvent_amt)

    req_amt = np.array(req_amt, dtype=float) + 2000
    req_amt[req_amt < 10000] = 10000
    req_amt = np.ceil(req_amt / 1000)

    out = solvent_map.copy()
    out['RequiredAmount'] = req_amt
    out = out.rename(columns={'Name': 'Name'})
    out['Labware'] = rack_position
    out['Unit'] = 'mL'
    out['Type'] = '50 mL Falcon Tube'
    out['Category'] = 'SOLVENT'
    out = out[['Category', 'Labware', 'Type', 'Slot', 'Name', 'RequiredAmount', 'Unit']]
    return out


def cal_stock_amt(sol_list, stock_list, stock_map, deck_map):
    amt_list = {k: 0.0 for k in stock_list}
    drugs = sol_list['DrugType'].unique().tolist()
    solvents = sol_list['Solvent'].unique().tolist()
    for drug in drugs:
        for solv in solvents:
            cur = sol_list[(sol_list['DrugType'] == drug) & (sol_list['Solvent'] == solv)]
            if len(cur) > 0:
                max_conc = cur['DrugConc'].astype(float).max()
                cur_req = cur.loc[cur['DrugConc'].astype(float) == max_conc, 'AmtHi'].astype(float).sum()
                amt_list[cur['DrugType'].iloc[0]] += cur_req

    rows = []
    for name, conc in stock_list.items():
        req = amt_list.get(name, 0.0) + 300
        req = max(req, 700)
        req = math.ceil(req / 100) * 100
        rows.append({'Name': name, 'Conc': conc, 'RequiredAmount': req})
    df = pd.DataFrame(rows)

    df = df.merge(stock_map[['Slot', 'Name']], on='Name', how='left')
    labware = _name_for_value(deck_map, '15ml_Falcon_stock')
    df['Labware'] = labware
    df['Unit'] = 'uL'
    df['Type'] = '15 mL Falcon Tube'
    df['Category'] = 'DRUG STOCK'
    df = df[['Category', 'Labware', 'Type', 'Slot', 'Name', 'RequiredAmount', 'Unit']]
    return df


def make_stock_map(deck_map, stock_list):
    coords = [0, 0]
    rows = []
    for name in stock_list:
        slot = f"{LETTERS[coords[0]]}{coords[1] + 1}"
        rows.append({'Slot': slot, 'Name': name})
        coords[1] += 1
        if coords[1] > 4:  # 15 mL Falcon rack: 5 per row
            coords[1] = 0
            coords[0] += 1
    return pd.DataFrame(rows)


def make_solvent_map(plate_map):
    coords = [0, 0]
    rows = []
    solvents = plate_map['Solvent'].unique().tolist()
    for solv in solvents:
        slot = f"{LETTERS[coords[0]]}{coords[1] + 1}"
        rows.append({'Slot': slot, 'Name': str(solv)})
        coords[1] += 1
        if coords[1] > 2:  # 50 mL Falcon rack: 3 per row
            coords[1] = 0
            coords[0] += 1
    return pd.DataFrame(rows)


def make_inoc_map(plate_map):
    coords = [0, 0]
    rows = []
    inocs = [i for i in plate_map['Inoc'].unique().tolist() if i != 'NA']
    for inoc in inocs:
        slot = f"{LETTERS[coords[0]]}{coords[1] + 1}"
        rows.append({'Slot': slot, 'Name': str(inoc)})
        coords[1] += 1
        if coords[1] > 4:
            coords[1] = 0
            coords[0] += 1
    return pd.DataFrame(rows)


def cal_dil_tubes(dil_map):
    # R's table() orders by the character labels' natural sort order
    # (factor levels default to sorted unique values), NOT by count --
    # e.g. "labware_10" < "labware_8" < "labware_9" as strings.
    occs = dil_map['Labware'].value_counts()
    occs = occs.reindex(sorted(occs.index))
    df = pd.DataFrame({
        'Category': 'EMPTY TUBES FOR DILUTION',
        'Labware': occs.index,
        'Type': '-',
        'Slot': '-',
        'Name': '15_Falcon',
        'RequiredAmount': occs.values,
        'Unit': 'tubes',
    })
    return df.reset_index(drop=True)


def cal_deck_adjustment(cmd_list, deck_map, dil_tubes, n_plate):
    deck_vals = list(range(12, 0, -1))
    deck = [deck_vals[i:i + 3] for i in range(0, 12, 3)]
    deck = [[row[2], row[1], row[0]] for row in deck]

    dm_vals = list(deck_map.values())
    deck_map_mat = [dm_vals[i:i + 3] for i in range(0, 12, 3)]

    fin_deck = []
    for i in range(4):
        fin_deck.append(list(deck[i]))
        fin_deck.append(list(deck_map_mat[3 - i]))

    if 'labware_8' not in dil_tubes['Labware'].values:
        fin_deck[3][1] = '(empty)'

    if n_plate < 6:
        fin_deck[7][2] = '(empty)'
        if n_plate < 5:
            fin_deck[7][1] = '(empty)'
            if n_plate < 4:
                fin_deck[7][0] = '(empty)'
                if n_plate < 3:
                    fin_deck[5][2] = '(empty)'
                    if n_plate < 2:
                        fin_deck[5][1] = '(empty)'
    return fin_deck


# ---------------------------------------------------------------------------
# ADJUSTMENTS
# ---------------------------------------------------------------------------

def convert_amtlist_mv_to_mc(amt_list):
    return pd.DataFrame({
        'Labware': amt_list['Labware'],
        'Slot': amt_list['Slot'],
        'Fill': amt_list['Name'],
        'Conc': 0,
        'Vol': amt_list['RequiredAmount'].astype(float),
    })


def cal_amtlist_excess(amt_list, cmd_list, deck_map2):
    """deck_map2: DataFrame with columns [Labware(key), Desc]"""
    tubes = amt_list.copy()
    tubes['Vol'] = 0.0
    cmd_list = cmd_list.copy()

    def n_targets(s):
        return len([w for w in str(s).split(', ') if w != ''])

    for i in range(len(cmd_list)):
        src_ware = cmd_list['SourceLabware'].iloc[i]
        src_slot = cmd_list['SourceSlot'].iloc[i]
        if src_ware in tubes['Labware'].values:
            mask_tube = (tubes['Labware'] == src_ware) & (tubes['Slot'] == src_slot)
            if mask_tube.any():
                delta_v = float(cmd_list['TransAmt'].iloc[i]) * n_targets(cmd_list['TargetSlot'].iloc[i])
                vol_used_after = tubes.loc[mask_tube, 'Vol'].iloc[0] + delta_v

                limit_ok = ((vol_used_after <= 45000 and src_ware == 'labware_11') or
                            (vol_used_after <= 14000 and src_ware == 'labware_10'))

                if limit_ok:
                    tubes.loc[mask_tube, 'Vol'] = tubes.loc[mask_tube, 'Vol'] + delta_v
                else:
                    sol_or_stock = 'olvent' if src_ware == 'labware_11' else 'tock'
                    desc_match = deck_map2[deck_map2['Desc'].str.contains(sol_or_stock)]
                    lab_id = desc_match['Labware'].iloc[0]
                    same_lab = tubes[tubes['Labware'] == lab_id]
                    last_filled = same_lab['Slot'].iloc[-1]

                    new_row_letter, new_col = last_filled[0], int(last_filled[1:])
                    new_col += 1
                    if new_col > 3:
                        new_row_letter = chr(ord(new_row_letter) + 1)
                        new_col = 1
                    new_slot = f"{new_row_letter}{new_col}"

                    fill_val = tubes.loc[mask_tube, 'Fill'].iloc[0]
                    conc_val = tubes.loc[mask_tube, 'Conc'].iloc[0]
                    new_row = pd.DataFrame([{
                        'Labware': src_ware, 'Slot': new_slot,
                        'Fill': fill_val, 'Conc': conc_val, 'Vol': delta_v,
                    }])
                    tubes = pd.concat([tubes, new_row], ignore_index=True)

                    for j in range(i, len(cmd_list)):
                        if (cmd_list['SourceLabware'].iloc[j] == src_ware and
                                cmd_list['SourceSlot'].iloc[j] == src_slot):
                            cmd_list.loc[j, 'SourceSlot'] = new_slot

    amt_list_out = tubes.sort_values('Labware', kind='stable').reset_index(drop=True)

    def clip_round(mask, div, add, minv, maxv):
        v = amt_list_out.loc[mask, 'Vol'].astype(float)
        v = np.ceil(v / div) + add if add and div == 1000 else v
        return v

    solvent_lab = deck_map2.loc[deck_map2['Desc'].str.contains('olvent'), 'Labware'].iloc[0]
    stock_lab = deck_map2.loc[deck_map2['Desc'].str.contains('tock'), 'Labware'].iloc[0]

    mask_solv = amt_list_out['Labware'] == solvent_lab
    v = amt_list_out.loc[mask_solv, 'Vol'].astype(float)
    v = np.ceil(v / 1000) + 3
    v = v.clip(lower=5, upper=48)
    amt_list_out.loc[mask_solv, 'Vol'] = v

    mask_stock = amt_list_out['Labware'] == stock_lab
    v = amt_list_out.loc[mask_stock, 'Vol'].astype(float)
    v = np.ceil(v / 100) * 100 + 300
    v = v.clip(lower=700, upper=14000)
    amt_list_out.loc[mask_stock, 'Vol'] = v

    return cmd_list, amt_list_out


def convert_amtlist_mc_to_mv(new_all_amt):
    category = new_all_amt['Labware'].apply(lambda x: 'SOLVENT' if x == 'labware_11' else 'STOCK')
    typ = new_all_amt['Labware'].apply(lambda x: '50 mL Falcon Tube' if x == 'labware_11' else '15 mL Falcon Tube')
    unit = new_all_amt['Labware'].apply(lambda x: 'mL' if x == 'labware_11' else 'uL')
    out = pd.DataFrame({
        'Category': category, 'Labware': new_all_amt['Labware'], 'Type': typ,
        'Slot': new_all_amt['Slot'], 'Name': new_all_amt['Fill'],
        'RequiredAmount': new_all_amt['Vol'], 'Unit': unit,
    })
    return out